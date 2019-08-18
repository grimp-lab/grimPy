# coding: utf-8

'''
This module implements an API for reading and writing IRIS data from and to the GRIMP template spreadsheet. The API exposes core functions from
iris_core.py
'''

import pandas as pd
import xarray as xr
from openpyxl import load_workbook
from grimpy.core.instruments import iris_core as iris
from grimpy.utils.list utils import nanify_list
from grimpy.constants.instrumental_constants import IRIS_VERSIONS


def read_iris_data(filename):
    '''
    Reads IRIS data from the GRIMP spreadsheet template. Outputs a Python dictionnary gathering data of the IRIS session.
    :param filename: path to the file
    :type filename: string
    '''
    try:
        data = pd.read_excel(filename, sheet_name="IRIS")
    except :
        print('Unable to open file ' + filename + ' : ', sys.exc_info()[0])

    # IRIS version retrieval
    version = data["Version"].dropna().tolist()[0]
    if version not in IRIS_VERSIONS :
        raise Exception('Incorrect IRIS version specified : {}.'.format(version) + 'Choices are : ' + ' '.join(IRIS_VERSIONS))

    # Calibration measurements retrieval
    spectralons_reflectances = data["Spectralon (%)"].dropna().tolist()
    calibration_one = nanify_list(data["Calibration Voltage 1 (V)"].dropna().tolist())
    calibration_two = nanify_list(data["Calibration Voltage 2 (V)"].dropna().tolist())
    calibration_three = nanify_list(data["Calibration Voltage 3 (V)"].dropna().tolist())
    calibration_list = [calibration_one, calibration_two, calibration_three]
    # Convert each of the calibration lists into a numpy array
    calibration_scans = np.array([calib for calib in calibration_list if len(calib) > 0 ])
    calibration_voltages = np.nanmean(calibration_scans, axis=0)

    # Scans data retrieval
    samples_heights = data["Height (cm)"].dropna().tolist()
    scan_one = nanify_list(data["Scan 1 (V)"].dropna().tolist())
    scan_two = nanify_list(data["Scan 2 (V)"].dropna().tolist())
    scan_three = nanify_list(data["Scan 3 (V)"].dropna().tolist())
    scans = np.array([scan_one, scan_two, scan_three])
    # Average all three measurements
    samples_voltages = np.nanmean(scans, axis=0)

    # Build the data dictionnary
    iris_session = {
        'version': version,
        'spectralons_reflectances' : spectralons_reflectances,
        'calibratation_voltages': calibration_voltages,
        'samples_heights' : samples_heights,
        'samples_voltages': measures
    }

    return iris_session


def write_iris_sheet(filename, iris_session):
    '''
    Writes an iris_session dictionnary to the specified GRIMP spreadsheet.
    :param filename: path to the file
    :type filename: string
    '''

    # Open spreadsheet
    try :
        workbook = load_workbook(filename)
        iris_sheet = workbook["IRIS"]
    except:
        print('Unable to open file ' + filename + ' : ', sys.exc_info()[0])

    for i in range(len(iris_session['samples_voltages'])):
        iris_sheet.cell(row = i+2, column = 10).value = iris_session['samples_reflectances'][i]
        iris_sheet.cell(row = i+2, column = 11).value = iris_session['samples_ssa'][i]
        iris_sheet.cell(row = i+2, column = 12).value = iris_session['samples_optical_radii'][i]

    try:
        workbook.save(filename)
    except:
        print('Unable to save file ' + filename + ' : ', sys.exc_info()[0])

    return('Success - spreadsheet correctly updated.')


def process_iris_session(iris_session):
    '''
    Processes a Python dictionnary in the shape of an iris_session defined in get_iris_data. Outpus a disctionnary with 3 new keys-list :
    samples_reflectances, samples_ssa, samples_optical_radii.
    :param iris_session: IRIS cleaned measurement session in the shape of an iris_session defined in read_iris_data
    :type iris_session: Python dictionnary
    '''

    # Check for input dictionnary legality
    if (iris_session.keys() != ['version', 'spectralons_reflectances', 'calibration_voltages', 'samples_heights', 'samples_voltages']) :
        raise('Input dictionnary is not a legal iris_session. Please refer to get_iris_data for more details on iris_session.')

    # Reflectance calculation
    reflectances = volt_to_reflectance(measures, polynom)

    # SSA calculation and optical radius calculation
    if version in IRIS_VERSIONS :
        ssa = reflectance_to_ssa(reflectances, version)
        r_opt = SSA_to_r_opt(ssa)
    else:
        raise Exception('Incorrect IRIS version specified : {}.'.format(version) + 'Choices are : ' + ' '.join(IRIS_VERSIONS))

    iris_session.update({
        'samples_reflectances': reflectances,
        'samples_ssa': ssa,
        'samples_optical_radii': r_opt
    })

    return iris_session
